"""Exportación a CSV, JSON y Excel. Genera los archivos en memoria."""
from __future__ import annotations

import csv
import io
import json
from typing import Any, Dict, List

from app.db.database import get_connection

COLUMNS = [
    ("product", "Producto"),
    ("game", "Juego"),
    ("set_name", "Set"),
    ("product_type_name", "Tipo"),
    ("language", "Idioma"),
    ("units_total", "Unidades"),
    ("store", "Tienda"),
    ("price", "Precio"),
    ("currency", "Moneda"),
    ("unit_price", "Precio por unidad"),
    ("stock", "Stock"),
    ("is_best_price", "Mejor precio"),
    ("match_method", "Método de agrupación"),
    ("match_score", "Confianza"),
    ("url", "URL"),
    ("last_seen_at", "Actualizado"),
]

_STOCK_LABEL = {
    "in_stock": "En stock",
    "out_of_stock": "Agotado",
    "preorder": "Preventa",
    "coming_soon": "Próximamente",
    "unknown": "Desconocido",
}


def _language_label(code: Any) -> str:
    from app.core import attributes as attrs_module

    return attrs_module.language_name(code) or "Sin declarar"


def collect_rows(product_ids: List[int] | None = None) -> List[Dict[str, Any]]:
    sql = """
        SELECT p.id AS product_id, p.display_name, p.game, p.set_name,
               p.product_type_name, p.units_total AS product_units,
               p.best_available_price,
               sp.price, sp.currency, sp.stock_status, sp.url, sp.last_seen_at,
               sp.units_total, sp.quantity_confidence, sp.language,
               sp.match_method, sp.match_score,
               s.name AS store_name
        FROM store_products sp
        JOIN products p ON p.id = sp.product_id
        JOIN stores s ON s.id = sp.store_id
        WHERE sp.is_active = 1
    """
    params: tuple = ()
    if product_ids:
        placeholders = ",".join("?" * len(product_ids))
        sql += f" AND p.id IN ({placeholders})"
        params = tuple(product_ids)
    sql += " ORDER BY p.display_name, sp.price"

    from app.core.units import unit_price

    with get_connection() as conn:
        rows = conn.execute(sql, params).fetchall()

    out: List[Dict[str, Any]] = []
    for row in rows:
        out.append(
            {
                "product": row["display_name"],
                "game": row["game"],
                "set_name": row["set_name"],
                "product_type_name": row["product_type_name"],
                "language": _language_label(row["language"]),
                "units_total": row["units_total"],
                "store": row["store_name"],
                "price": row["price"],
                "currency": row["currency"],
                "unit_price": unit_price(
                    row["price"], row["units_total"], float(row["quantity_confidence"] or 0)
                ),
                "stock": _STOCK_LABEL.get(row["stock_status"], row["stock_status"]),
                "is_best_price": "Sí"
                if row["best_available_price"] is not None
                and row["price"] == row["best_available_price"]
                else "No",
                "match_method": row["match_method"],
                "match_score": row["match_score"],
                "url": row["url"],
                "last_seen_at": row["last_seen_at"],
            }
        )
    return out


def to_csv(rows: List[Dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
    writer.writerow([label for _key, label in COLUMNS])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _label in COLUMNS])
    # BOM para que Excel en Windows abra los acentos correctamente.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_json(rows: List[Dict[str, Any]]) -> bytes:
    return json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")


def to_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comparación"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for index, (_key, label) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        for col_index, (key, _label) in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_index, column=col_index, value=row.get(key))

    widths = {"product": 46, "url": 52, "set_name": 22, "product_type_name": 20, "store": 18}
    for index, (key, label) in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(key, max(12, len(label) + 2))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, len(rows) + 1)}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
